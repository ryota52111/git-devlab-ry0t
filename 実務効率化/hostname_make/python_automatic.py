import openpyxl
import glob
import os

save_name1 = []
book_name = "./result.xlsx"

files = glob.glob("./tmp/*") #tmp内のファイルを取得
for file in files: #取得したファイルをforで回す。
    fail = os.path.basename(file)
    save_name1.append(fail) #tmp内のファイルをtextに格納

def read_file():
    for i, v in enumerate(save_name1):    
        with open(f"./tmp/{v}", encoding='shift-jis') as f:
            lines = f.read().splitlines()
            wb = openpyxl.load_workbook(book_name)
            crete_sheet = wb.create_sheet(v)
            sheet = wb[v]
            for i, v in enumerate(lines):
                sheet.cell(row=1 + i, column=1, value=v)
            wb.save(book_name)



read_file()



